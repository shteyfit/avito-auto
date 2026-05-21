# -*- coding: utf-8 -*-
"""
Программа расчёта итоговых цен по остаткам:
- Kolobox (комиссия + каталог)
- 4tochki (ОХ-4) по файлу Forto4ki_tires_detailed.xlsx (листы car, vned, cartruck)
- Статус "Купленный"
- Ручные фиксы из fixprice.xlsx

Требуемые библиотеки:
    pip install pandas openpyxl

Требуемые файлы в папке (или указать пути в интерфейсе):
    result_with_articul.xlsx
    kolobox_commission_all_points.xlsx
    catalog_tyres.xlsx
    Forto4ki_tires_detailed.xlsx
    fixprice.xlsx (опционально)

Особенности:
- Размер шин берётся из структурных колонок:
  * kolobox_commission_all_points: tyres_tread_width / tyres_profile_height / tyres_diameter
  * catalog_tyres: tread_width / profile_height / diameter
  * Forto4ki_tires_detailed (листы car/vned/cartruck): Ширина / Профиль / Диаметр
- Сезон определяется по строке-разделителю "Зимние шины" в колонке 'Название товара':
  * всё выше — летние
  * всё ниже — зимние
- Если нельзя подобрать цену в заданных рамках маржи (слишком большой разброс закупочных цен),
  цена оставляется пустой, товар помечается "Нужна ручная цена", строка подсвечивается.
"""

import os
import math
import re
import datetime as dt
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd


# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

def parse_size_from_title(title: str):
    """Резервный парсинг размера из строки вида '205/55 R16', '205/5516' → '205/55R16'."""
    if not isinstance(title, str):
        return None
    m = re.search(r'(\d{3})\s*/\s*(\d{2})\s*R?\s*(\d{2})', title)
    if not m:
        return None
    w, p, d = m.groups()
    return f"{w}/{p}R{d}"


def determine_seasons(df_result, log_func=None):
    """
    Определяем сезон по строке-разделителю "Зимние шины" в колонке 'Название товара':
    - всё выше неё: summer
    - всё ниже: winter
    Саму строку-разделитель удаляем.
    """
    def _log(msg):
        if log_func:
            log_func(msg)

    df = df_result.copy()
    col = 'Название товара'
    if col not in df.columns:
        raise ValueError("В файле result_with_articul нет колонки 'Название товара'.")

    mask = df[col].astype(str).str.contains('Зимние шины', case=False, na=False)
    idxs = df.index[mask].tolist()
    if not idxs:
        _log("ОШИБКА: не найдена строка 'Зимние шины' в колонке 'Название товара'.")
        raise ValueError("Не найдена строка 'Зимние шины' — разделите файл на летние/зимние шины.")

    sep_idx = idxs[0]
    _log(f"Найдена строка 'Зимние шины' на индексе {sep_idx}.")

    seasons = []
    for idx in df.index:
        if idx < sep_idx:
            seasons.append('summer')
        elif idx > sep_idx:
            seasons.append('winter')
        else:
            seasons.append('separator')

    df['__season'] = seasons
    df = df[df['__season'] != 'separator'].copy()
    df.reset_index(drop=True, inplace=True)
    return df


def load_forto4ki_detailed(path, sheets=('car', 'vned', 'cartruck'), log_func=None):
    """
    Загружаем Forto4ki_tires_detailed.xlsx:
    - берём только указанные листы (по умолчанию car, vned, cartruck)
    - склеиваем их в один DataFrame
    Ожидается, что на каждом листе первая строка — уже заголовки (Код товара, Марка, ..., Цена (первая) и т.п.).
    """
    def _log(msg):
        if log_func:
            log_func(msg)

    _log(f"Читаю Forto4ki_tires_detailed из листов: {', '.join(sheets)}")

    xls = pd.ExcelFile(path)
    frames = []
    for sheet in sheets:
        if sheet in xls.sheet_names:
            _log(f"  - лист '{sheet}' найден, читаю...")
            df_sheet = pd.read_excel(xls, sheet_name=sheet)
            frames.append(df_sheet)
        else:
            _log(f"  - лист '{sheet}' НЕ найден, пропускаю.")

    if not frames:
        _log("Не удалось загрузить ни одного листа из Forto4ki_tires_detailed.")
        return pd.DataFrame()

    df_all = pd.concat(frames, ignore_index=True)
    _log(f"Forto4ki: объединённый размер: {df_all.shape[0]} строк, {df_all.shape[1]} столбцов.")
    return df_all


def build_size_map(commission_df, catalog_df, fort_df):
    """
    Строим словарь: articul (строка Id) -> canonical size '205/55R16'.
    Приоритет источников:
      1) kolobox_commission_all_points (tyres_tread_width/profile_height/diameter)
      2) catalog_tyres (tread_width/profile_height/diameter)
      3) Forto4ki_tires_detailed (Ширина/Профиль/Диаметр, Код товара)
    """
    size_map = {}

    # 1) Kolobox commission
    if commission_df is not None and 'articul' in commission_df.columns:
        df = commission_df.copy()
        df['articul'] = df['articul'].astype(str)
        for _, row in df.iterrows():
            art = row['articul']
            if art in size_map:
                continue
            w = row.get('tyres_tread_width')
            h = row.get('tyres_profile_height')
            d = row.get('tyres_diameter')
            if pd.isna(w) or pd.isna(h) or pd.isna(d):
                continue
            try:
                w = int(float(w))
                h = int(float(h))
                d = int(str(d).replace(',', '.'))
            except Exception:
                continue
            size_map[art] = f"{w}/{h}R{d}"

    # 2) catalog_tyres
    if catalog_df is not None and 'articul' in catalog_df.columns:
        df = catalog_df.copy()
        df['articul'] = df['articul'].astype(str)
        for _, row in df.iterrows():
            art = row['articul']
            if art in size_map:
                continue
            w = row.get('tread_width')
            h = row.get('profile_height')
            d = row.get('diameter')
            if pd.isna(w) or pd.isna(h) or pd.isna(d):
                continue
            try:
                w = int(float(w))
                h = int(float(h))
                d = int(str(d).replace(',', '.'))
            except Exception:
                continue
            size_map[art] = f"{w}/{h}R{d}"

    # 3) Forto4ki_tires_detailed (объединённый df)
    if fort_df is not None and not fort_df.empty:
        df = fort_df.copy()
        cols = {str(c).strip().lower(): c for c in df.columns}
        code_col = cols.get('код товара') or cols.get('код')
        w_col = cols.get('ширина')
        h_col = cols.get('профиль')
        d_col = cols.get('диаметр')

        if code_col and w_col and h_col and d_col:
            df[code_col] = df[code_col].astype(str)
            for _, row in df.iterrows():
                art = row[code_col]
                if art in size_map:
                    continue
                w = row.get(w_col)
                h = row.get(h_col)
                d = row.get(d_col)
                if pd.isna(w) or pd.isna(h) or pd.isna(d):
                    continue
                try:
                    w = int(float(w))
                    h = int(float(h))
                    d = int(float(d))
                except Exception:
                    continue
                size_map[art] = f"{w}/{h}R{d}"

    return size_map


def build_size_popularity(df_result, size_map, popular_threshold=3, rare_threshold=2):
    """
    Определяем популярность размеров:
      - 'popular' если вариантов в размере ≥ popular_threshold
      - 'rare'    если ≤ rare_threshold
      - 'normal'  иначе

    Размер берём в первую очередь из size_map[Id],
    и только если там None — парсим из 'Название товара'.
    """
    df = df_result.copy()
    df['Id'] = df['Id'].astype(str)

    df['__size_key'] = df['Id'].map(size_map)

    mask_missing = df['__size_key'].isna()
    if mask_missing.any():
        df.loc[mask_missing, '__size_key'] = df.loc[mask_missing, 'Название товара'].apply(
            parse_size_from_title
        )

    size_counts = df['__size_key'].value_counts(dropna=True).to_dict()

    size_pop = {}
    for size, count in size_counts.items():
        if count >= popular_threshold:
            size_pop[size] = 'popular'
        elif count <= rare_threshold:
            size_pop[size] = 'rare'
        else:
            size_pop[size] = 'normal'

    return df, size_pop


def compute_kolobox_distribution(df_commission, address_ids):
    """
    Строим карту: articul -> список партий (price, quantity)
    только по нужным складам.
    """
    df = df_commission[df_commission['address_id'].isin(address_ids)].copy()
    df = df[df['articul'].notna() & df['price'].notna()]
    df['articul'] = df['articul'].astype(str)
    df['price'] = pd.to_numeric(df['price'], errors='coerce')
    df['quantity'] = pd.to_numeric(df['quantity'], errors='coerce').fillna(0)
    df = df[df['quantity'] > 0]

    grouped = {}
    for articul, g in df.groupby('articul'):
        g2 = g.groupby('price', as_index=False)['quantity'].sum()
        batches = [(row['price'], row['quantity']) for _, row in g2.iterrows()]
        grouped[articul] = batches
    return grouped


def compute_catalog_price_map(catalog_df):
    cat = catalog_df.copy()
    cat['articul'] = cat['articul'].astype(str)
    cat['price'] = pd.to_numeric(cat['price'], errors='coerce')
    return cat.set_index('articul')['price'].to_dict()


def compute_forto4ki_price_map(fort_df):
    """
    Карта: Код товара -> Цена (первая) для объединённого df из Forto4ki_tires_detailed.
    Ожидаемые колонки: 'Код товара', 'Цена (первая)'.
    """
    if fort_df is None or len(fort_df) == 0:
        return {}

    df = fort_df.copy()
    cols = {str(c).strip().lower(): c for c in df.columns}
    code_col = cols.get('код товара') or cols.get('код')
    price_col = None
    for key, orig in cols.items():
        if ('цена' in key) and ('первая' in key) and ('розн' not in key):
            price_col = orig
            break

    if code_col is None or price_col is None:
        return {}

    df[code_col] = df[code_col].astype(str)
    df[price_col] = pd.to_numeric(df[price_col], errors='coerce')
    return df.set_index(code_col)[price_col].to_dict()


def compute_fixprice_map(df_fix):
    if df_fix is None:
        return {}
    df = df_fix.copy()
    cols = {str(c).lower(): c for c in df.columns}
    art_col = cols.get('артикул') or cols.get('id') or list(df.columns)[0]
    price_col = cols.get('цена') or list(df.columns)[1]
    df[art_col] = df[art_col].astype(str)
    df[price_col] = pd.to_numeric(df[price_col], errors='coerce')
    return df.set_index(art_col)[price_col].to_dict()


def choose_margin_range(qty, global_min, global_max, low_qty=4, high_qty=20):
    """
    Диапазон маржи в зависимости от количества:
    - мало штук → ближе к верхнему краю
    - много штук → ближе к нижнему
    """
    rng = global_max - global_min
    if qty is None or (isinstance(qty, float) and math.isnan(qty)):
        qty = low_qty

    if qty <= low_qty:
        low = global_min + 0.5 * rng
        high = global_max
    elif qty >= high_qty:
        low = global_min
        high = global_min + 0.5 * rng
    else:
        t = (qty - low_qty) / (high_qty - low_qty)
        low_low = global_min + 0.5 * rng
        low_high = global_max
        high_low = global_min
        high_high = global_min + 0.5 * rng
        low = low_low + (high_low - low_low) * t
        high = low_high + (high_high - low_high) * t

    return low, high


def compute_effective_cost(batches, catalog_price=None, alpha_max=0.7):
    """
    batches: [(price, qty), ...]
    catalog_price: если меньше минимальной цены партии — считаем, что расчёт
    будет только по этой цене (акция), а комиссии игнорируем.

    Возвращаем: (эффективная себестоимость, минимальная цена, максимальная цена)
    """
    if not batches:
        return None, None, None

    prices = [p for p, q in batches]
    qtys = [q for p, q in batches]
    total_qty = sum(qtys)

    min_comm = min(prices)
    max_comm = max(prices)

    # Если цена из каталога ниже любой комиссионной — считаем только по каталогу.
    if catalog_price is not None and catalog_price > 0 and catalog_price < min_comm:
        eff = catalog_price
        return eff, catalog_price, catalog_price

    # Обычный режим: работаем только по комиссиям
    min_ref = min_comm

    if total_qty <= 0:
        eff = max_comm
    else:
        qty_min = sum(q for p, q in batches if p == min_comm)
        share = qty_min / total_qty  # доля самой дешёвой партии
        eff = max_comm - alpha_max * share * (max_comm - min_ref)

    return eff, min_ref, max_comm


def parse_keywords(text: str):
    """Разбираем ключевые слова из многострочного текста/через запятую."""
    if not text:
        return []
    parts = re.split(r'[,;\n\r]+', text)
    return [p.strip().lower() for p in parts if p.strip()]


def file_age_days(path):
    ts = os.path.getmtime(path)
    return (dt.datetime.now() - dt.datetime.fromtimestamp(ts)).days


def color_special_rows_in_excel(path, log_func=None):
    """Подсветить строки со статусом 'Купленный' и 'Нужна ручная цена'."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
    except ImportError:
        if log_func:
            log_func("openpyxl не установлен — пропускаю раскраску.")
        return

    if log_func:
        log_func("Раскрашиваем строки 'Купленный' и 'Нужна ручная цена'...")

    wb = load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    try:
        status_idx = headers.index("Статус")
    except ValueError:
        status_idx = None

    try:
        manual_idx = headers.index("Нужна ручная цена")
    except ValueError:
        manual_idx = None

    fill_kup = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    fill_manual = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        is_kup = False
        is_manual = False

        if status_idx is not None:
            val = row[status_idx].value
            if isinstance(val, str) and val.strip().lower() == "купленный":
                is_kup = True

        if manual_idx is not None:
            val2 = row[manual_idx].value
            if isinstance(val2, (bool, str, int, float)):
                if str(val2).strip().lower() in ("true", "1", "да"):
                    is_manual = True

        if is_kup:
            for cell in row:
                cell.fill = fill_kup
        elif is_manual:
            for cell in row:
                cell.fill = fill_manual

    wb.save(path)
    if log_func:
        log_func("Раскраска завершена.")


# ==================== ОСНОВНОЙ РАСЧЁТ ====================

def compute_auto_prices(
    df_result,
    commission_df,
    catalog_df,
    fort_df,
    fixprice_df=None,
    address_ids=(229, 2139, 2138, 4344),
    global_min_margin=0.06,
    global_max_margin=0.13,
    low_qty=4,
    high_qty=20,
    popular_threshold=3,
    rare_threshold=2,
    popular_shift=-0.01,
    rare_shift=0.01,
    alpha_max=0.7,
    ox4_margin_default=0.10,
    summer_brand_keys=None,
    summer_model_keys=None,
    winter_brand_keys=None,
    winter_model_keys=None,
    summer_brand_keys_unpop=None,
    summer_model_keys_unpop=None,
    winter_brand_keys_unpop=None,
    winter_models_unpop=None,
    log_func=None,
):
    """
    Основная функция расчёта.
    """

    def _log(msg):
        if log_func:
            log_func(msg)

    summer_brand_keys = summer_brand_keys or []
    summer_model_keys = summer_model_keys or []
    winter_brand_keys = winter_brand_keys or []
    winter_model_keys = winter_model_keys or []

    summer_brand_keys_unpop = summer_brand_keys_unpop or []
    summer_model_keys_unpop = summer_model_keys_unpop or []
    winter_brand_keys_unpop = winter_brand_keys_unpop or []
    winter_models_unpop = winter_models_unpop or []

    _log("Строим карту размеров по артикулам...")
    size_map = build_size_map(commission_df, catalog_df, fort_df)

    _log("Строим карту размеров и их популярности...")
    size_df, size_pop = build_size_popularity(df_result, size_map,
                                              popular_threshold, rare_threshold)

    # Карта статусов по Id
    status_map = {}
    for _, r in size_df.iterrows():
        rid = str(r['Id'])
        st = str(r['Статус']).strip().lower()
        status_map.setdefault(rid, set()).add(st)

    arts_with_commission = {
        rid for rid, sts in status_map.items()
        if any('комис' in s for s in sts)
    }
    arts_with_kup = {
        rid for rid, sts in status_map.items()
        if any('куплен' in s for s in sts)
    }

    _log("Строим распределение партий комиссии по складам...")
    comm_map = compute_kolobox_distribution(commission_df, address_ids)

    _log("Строим карту цен каталога Kolobox...")
    cat_map = compute_catalog_price_map(catalog_df)

    _log("Строим карту цен Forto4ki (ОХ-4)...")
    fort_map = compute_forto4ki_price_map(fort_df)

    # Загружаем fixprice
    fix_map = compute_fixprice_map(fixprice_df) if fixprice_df is not None else {}

    if fix_map:
        _log(f"Загружено {len(fix_map)} фиксов цен из fixprice.")

    # --- Логика конфликтов fixprice ---
    # Конфликт: артикул есть в fixprice и одновременно есть в остатках со статусом "На комиссию".
    # Каталоги/комиссионные прайсы здесь не учитываем.
    fix_conflict_ids = set()

    if fix_map:
        for fid in fix_map.keys():
            # учитываем обе формы Id: с нулём и без нуля
            variants = {fid}
            if fid.startswith("0"):
                variants.add(fid.lstrip("0"))
            else:
                variants.add("0" + fid)

            # конфликт, если этот артикул присутствует среди комиссионных Id в остатках
            has_commission_here = any(v in arts_with_commission for v in variants)

            if has_commission_here:
                fix_conflict_ids.add(fid)

        if fix_conflict_ids:
            _log(
                "Внимание: следующие артикула есть в fixprice и одновременно "
                "есть в остатках со статусом 'На комиссию' — для них будет "
                "использована цена Kolobox, а записи в fixprice рекомендуется "
                "проверить/удалить:\n"
                + ", ".join(sorted(fix_conflict_ids))
            )

    # -----------------------------------------------------

    out_rows = []
    manual_needed_count = 0

    for idx, row in size_df.iterrows():
        raw_id = str(row['Id'])
        status = str(row['Статус'])
        name = row['Название товара']
        qty_total = row['Количество']
        season = row.get('__season', 'summer')

        # Если товар и на комиссии, и купленный — купленный игнорируем
        if raw_id in arts_with_commission and status.strip().lower() == 'купленный':
            continue

        # Восстанавливаем ведущий ноль
        art = raw_id
        if art not in comm_map and art not in cat_map and art not in fort_map:
            art0 = "0" + art
            if art0 in comm_map or art0 in cat_map or art0 in fort_map:
                art = art0

        base_source = None
        base_cost = None
        min_cost = None
        max_cost = None
        final_price = None
        note = ""
        needs_manual = False
        size_key = row['__size_key']
        is_ox4 = '(ОХ-4' in str(name)

        # -----------------------------
        # 1) FIXPRICE — НАИВЫСШИЙ ПРИОРИТЕТ
        #    если нет комиссионной позиции в остатках
        # -----------------------------
        fix_key = None
        if raw_id in fix_map:
            fix_key = raw_id
        elif art in fix_map:
            fix_key = art

        if fix_key is not None:
            if fix_key not in fix_conflict_ids:
                final_price = float(fix_map[fix_key])
                base_source = 'fixprice'
                note = "Цена взята из fixprice.xlsx"
            # если в конфликте — fixprice игнорируется, идём дальше

        # 2) Купленный (если ещё не проставили fixprice)
        if (final_price is None) and (status.strip().lower() == 'купленный'):
            base_source = 'kup'
            needs_manual = True
            note = "Статус 'Купленный' — цена не рассчитывалась"

        # 3) OX-4
        elif (final_price is None) and is_ox4:
            base_source = '4tochki'
            base_cost = fort_map.get(art)
            if base_cost is None or base_cost <= 0:
                needs_manual = True
                note = "ОХ-4: не найдена цена"
            else:
                final_price = base_cost * (1 + ox4_margin_default)
                min_cost = max_cost = base_cost
                note = f"ОХ-4: базовая {base_cost:.2f}, маржа {ox4_margin_default*100:.1f}%"

        # 4) Kolobox: комиссия / каталог
        elif final_price is None:
            base_source = 'kolobox'
            batches = comm_map.get(art)
            cat_price = cat_map.get(art)
            used_catalog_only = False

            if not batches and not cat_price:
                needs_manual = True
                note = "Kolobox: нет комиссий и нет каталога"
            else:
                if batches:
                    base_cost, min_cost, max_cost = compute_effective_cost(
                        batches, catalog_price=cat_price, alpha_max=alpha_max
                    )
                else:
                    base_cost = cat_price
                    min_cost = max_cost = cat_price
                    used_catalog_only = True

                margin_low, margin_high = choose_margin_range(
                    qty_total, global_min_margin, global_max_margin, low_qty, high_qty
                )

                # Популярность размера
                if size_key in size_pop:
                    if size_pop[size_key] == 'popular':
                        margin_low += popular_shift
                        margin_high += popular_shift
                    elif size_pop[size_key] == 'rare':
                        margin_low += rare_shift
                        margin_high += rare_shift

                # Популярные / непопулярные бренды и модели
                name_lower = name.lower()
                brand_model_shift = 0.0

                if season == 'summer':
                    if any(k in name_lower for k in summer_brand_keys):
                        brand_model_shift += popular_shift
                    if any(k in name_lower for k in summer_model_keys):
                        brand_model_shift += popular_shift
                    if any(k in name_lower for k in summer_brand_keys_unpop):
                        brand_model_shift += rare_shift
                    if any(k in name_lower for k in summer_model_keys_unpop):
                        brand_model_shift += rare_shift
                else:  # winter
                    if any(k in name_lower for k in winter_brand_keys):
                        brand_model_shift += popular_shift
                    if any(k in name_lower for k in winter_model_keys):
                        brand_model_shift += популяр_shift
                    if any(k in name_lower for k in winter_brand_keys_unpop):
                        brand_model_shift += rare_shift
                    if any(k in name_lower for k in winter_models_unpop):
                        brand_model_shift += rare_shift

                margin_low += brand_model_shift
                margin_high += brand_model_shift

                margin_low = max(global_min_margin, margin_low)
                margin_high = min(global_max_margin, margin_high)

                target_margin = (margin_low + margin_high) / 2

                if not base_cost:
                    needs_manual = True
                    note = "Kolobox: не удалось вычислить себестоимость"
                else:
                    P0 = base_cost * (1 + target_margin)

                    c_min = min_cost if min_cost else base_cost
                    c_max = max_cost if max_cost else base_cost

                    min_allowed = c_max * (global_min_margin + 1)
                    max_allowed = c_min * (global_max_margin + 1)

                    if max_allowed < min_allowed:
                        needs_manual = True
                        note = (
                            "Kolobox: слишком большой разброс закупочных цен — "
                            "невозможно подобрать цену в пределах заданной маржи."
                        )
                    else:
                        P = max(P0, min_allowed)
                        P = min(P, max_allowed)
                        final_price = P

                        # --- ВОЗВРАЩАЕМ ПОДРОБНЫЕ КОММЕНТАРИИ ---
                        if used_catalog_only:
                            note = (
                                "Kolobox: комиссионных партий нет, базовая себестоимость "
                                f"взята из catalog_tyres ({base_cost:.2f}); целевая маржа "
                                f"{target_margin*100:.1f}% (диапазон "
                                f"{margin_low*100:.1f}–{margin_high*100:.1f}%)."
                            )
                        else:
                            note = (
                                "Kolobox: базовая себестоимость рассчитана по партиям комиссии "
                                f"({base_cost:.2f}); целевая маржа {target_margin*100:.1f}% "
                                f"(диапазон {margin_low*100:.1f}–{margin_high*100:.1f}%)."
                            )

        # Округление итоговой цены (кроме fixprice)
        if final_price is not None and base_source != 'fixprice':
            final_price = math.ceil(final_price / 10.0) * 10

        if needs_manual:
            manual_needed_count += 1

        out_row = dict(row)
        out_row.pop('__size_key', None)
        out_row['Id'] = art
        out_row['Источник'] = base_source
        out_row['Базовая себестоимость'] = base_cost
        out_row['Мин. закупочная'] = min_cost
        out_row['Макс. закупочная'] = max_cost
        out_row['Итоговая цена'] = final_price
        out_row['Комментарий'] = note
        out_row['Нужна ручная цена'] = needs_manual

        out_rows.append(out_row)

    df_out = pd.DataFrame(out_rows)
    return df_out, manual_needed_count


# ==================== GUI-ПРИЛОЖЕНИЕ ====================

class PricingApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Расчёт цен по остаткам (Kolobox + 4tochki)")
        self.geometry("1200x780")

        self._build_ui()

    def log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        for r in range(7):
            self.rowconfigure(r, weight=0)
        self.rowconfigure(6, weight=1)

        top_frame = ttk.Frame(self)
        top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        top_frame.columnconfigure(0, weight=1)

        title_lbl = ttk.Label(
            top_frame,
            text="Программа расчёта цен по остаткам (Kolobox + 4tochki)",
            font=("Segoe UI", 12, "bold")
        )
        title_lbl.grid(row=0, column=0, sticky="w")

        self.instr_visible = tk.BooleanVar(value=False)
        instr_btn = tk.Button(
            top_frame,
            text="Изучить инструкцию",
            command=self.toggle_instr,
            bg="#ffcc80",
            activebackground="#ffb74d"
        )
        instr_btn.grid(row=0, column=1, sticky="e", padx=5)

        self.instr_frame = ttk.Frame(self, relief="groove", borderwidth=1)
        self.instr_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 5))
        self.instr_frame.columnconfigure(0, weight=1)
        self.instr_frame.rowconfigure(0, weight=1)

        instr_text = (
            "Инструкция (кратко):\n"
            "1. Убедитесь, что в файле result_with_articul.xlsx есть строка 'Зимние шины' в колонке 'Название товара':\n"
            "   - всё выше — летние шины, всё ниже — зимние.\n"
            "2. В папке должны быть файлы (или задайте пути вручную):\n"
            "   - result_with_articul.xlsx — основной остаток (Id, Название товара, Статус, Количество и т.п.).\n"
            "   - kolobox_commission_all_points.xlsx — выгрузка комиссии Kolobox (articul, price, quantity, address_id,...).\n"
            "   - catalog_tyres.xlsx — каталог шин Kolobox с колонкой price.\n"
            "   - Forto4ki_tires_detailed.xlsx — прайс 4tochki (берутся листы car, vned, cartruck).\n"
            "   - fixprice.xlsx (опционально) — 2 колонки: Артикул, Цена (ручные фиксы).\n"
            "3. При необходимости измените параметры маржи и ключи популярных/непопулярных брендов/моделей.\n"
            "4. Нажмите 'Запустить расчёт'. Программа создаст выходной файл (по умолчанию output_prices.xlsx).\n"
            "5. Статус 'Купленный' и товары, где нельзя подобрать цену в заданных рамках маржи,\n"
            "   помечаются как 'Нужна ручная цена' и подсвечиваются в Excel.\n"
            "\n"
            "Подробное описание работы программы:\n"
            "1. Чтение файлов и подготовка данных:\n"
            "   • result_with_articul.xlsx — основной список остатков. По строке 'Зимние шины' файл делится\n"
            "     на летние и зимние позиции. Для каждой строки запоминается Id, Название, Статус, Количество.\n"
            "   • kolobox_commission_all_points.xlsx — по указанным ID складов берутся все партии комиссии\n"
            "     (articul, price, quantity). Для каждого артикула строится список партий (цена, количество).\n"
            "   • catalog_tyres.xlsx — для каждого артикула берётся каталожная цена Kolobox (столбец price).\n"
            "   • Forto4ki_tires_detailed.xlsx — с листов car, vned, cartruck берутся 'Код товара' и 'Цена (первая)'.\n"
            "   • fixprice.xlsx — справочник ручных фиксированных цен (Артикул, Цена).\n"
            "   • При необходимости артикулам автоматически восстанавливается ведущий ноль (если '1234567890'\n"
            "     не найден, программа пробует '01234567890').\n"
            "\n"
            "2. Определение размеров и их популярности:\n"
            "   • Для каждого артикула строится размер вида '205/55R16' по данным Kolobox, catalog_tyres или 4tochki.\n"
            "   • Если размер из структурных колонок определить нельзя, программа пытается вытащить его\n"
            "     из названия (пример: '205/55 R16 ...').\n"
            "   • Считается, сколько разных товаров у нас в каждом размере. Если вариантов в размере ≥ порога\n"
            "     'Популярный размер', размер считается популярным; если ≤ порога 'Редкий размер' — редким.\n"
            "   • Для популярных размеров маржа сдвигается вниз (товар конкурентный), для редких — вверх.\n"
            "\n"
            "3. Приоритет источников цены для каждой строки:\n"
            "   • Если по одному Id есть и строки 'На комиссию', и 'Купленный', строки 'Купленный' полностью\n"
            "     игнорируются — товар считается комиссионным.\n"
            "   • Если для Id есть запись в fixprice, и этот Id НЕ стоит на комиссии, используется цена из fixprice.\n"
            "   • Если Id есть и в fixprice, и на комиссии, программа использует комиссию, а в лог пишет предупреждение\n"
            "     с перечнем таких артикулов.\n"
            "   • Для шин с пометкой '(ОХ-4)' цена берётся из прайса 4tochki и к ней добавляется заданная маржа.\n"
            "   • Для остальных шин используется Kolobox (комиссия + каталог):\n"
            "       - если есть комиссионные партии, строится эффективная себестоимость с учётом цен и количеств.\n"
            "       - если каталоговая цена ниже всех комиссионных цен, считается, что действует акция, и расчёт\n"
            "         идёт только от каталожной цены (комиссионные партии игнорируются).\n"
            "       - если комиссий нет, но есть цена в catalog_tyres — расчёт идёт только от неё.\n"
            "       - если нет ни комиссий, ни каталожной цены — нужна ручная цена.\n"
            "\n"
            "4. Расчёт маржи и итоговой цены:\n"
            "   • Задан глобальный диапазон маржи (min–max). В зависимости от количества на складе\n"
            "     диапазон сдвигается: для небольших остатков маржа ближе к верхней границе,\n"
            "     для крупных остатков — ближе к нижней.\n"
            "   • Далее на диапазон влияют:\n"
            "       - популярность размера (популярный/редкий),\n"
            "       - популярность/непопулярность бренда и модели (по введённым ключевым словам отдельно\n"
            "         для лета и зимы).\n"
            "   • Итоговая целевая маржа берётся как середина скорректированного диапазона.\n"
            "   • Считается черновая цена: Себестоимость * (1 + целевая маржа).\n"
            "   • Дополнительно проверяется, чтобы итоговая цена не выходила за рамки маржи\n"
            "     относительно минимальной и максимальной закупочной цены по партиям.\n"
            "   • Если в рамках заданной маржи цену подобрать нельзя (слишком большой разброс закупочных цен),\n"
            "     товар отмечается как 'Нужна ручная цена'.\n"
            "   • Для всех автоматических цен (кроме fixprice) итоговая цена округляется вверх до ближайших 10 руб.\n"
            "\n"
            "5. Работа с fixprice:\n"
            "   • Если Id есть в fixprice и нет на комиссии — цена берётся строго из fixprice.\n"
            "   • Если Id есть и в fixprice, и на комиссии — используется только цена комиссии, а в лог\n"
            "     выводится список таких артикулов.\n"
            "   • В конце расчёта программа проверяет, какие артикула остались в fixprice, но отсутствуют\n"
            "     в текущем остатке — они выводятся в лог как кандидаты на удаление из файла fixprice.\n"
            "\n"
            "6. Итоговый файл:\n"
            "   • Для каждой строки сохраняются: исходные данные, источник цены (Kolobox, 4tochki, fixprice, Купленный),\n"
            "     базовая себестоимость, минимальная и максимальная закупочная цена, итоговая цена и комментарий.\n"
            "   • Строки со статусом 'Купленный' и строки, где 'Нужна ручная цена = Да', подсвечиваются\n"
            "     разными цветами в Excel для удобства последующей работы.\n"
        )
        self.instr_text_widget = tk.Text(self.instr_frame, height=15, wrap="word")
        self.instr_text_widget.insert("1.0", instr_text)
        self.instr_text_widget.configure(state="disabled")
        self.instr_text_widget.grid(row=0, column=0, sticky="nsew")
        instr_scroll = ttk.Scrollbar(self.instr_frame, command=self.instr_text_widget.yview)
        instr_scroll.grid(row=0, column=1, sticky="ns")
        self.instr_text_widget.configure(yscrollcommand=instr_scroll.set)

        self.instr_frame.grid_remove()

        files_frame = ttk.LabelFrame(self, text="Файлы")
        files_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        for c in range(3):
            files_frame.columnconfigure(c, weight=(1 if c == 1 else 0))

        self.path_result = tk.StringVar(value="result_with_articul.xlsx")
        self.path_commission = tk.StringVar(value="kolobox_commission_all_points.xlsx")
        self.path_catalog = tk.StringVar(value="catalog_tyres.xlsx")
        self.path_forto4ki = tk.StringVar(value="Forto4ki_tires_detailed.xlsx")
        self.path_fixprice = tk.StringVar(value="fixprice.xlsx")
        self.path_output = tk.StringVar(value="output_prices.xlsx")

        def add_file_row(row, label_text, var):
            ttk.Label(files_frame, text=label_text).grid(row=row, column=0, sticky="w")
            entry = ttk.Entry(files_frame, textvariable=var)
            entry.grid(row=row, column=1, sticky="ew", padx=5, pady=2)

            def browse():
                path = filedialog.askopenfilename(
                    title=f"Выберите файл: {label_text}",
                    filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
                )
                if path:
                    var.set(path)

            btn = tk.Button(
                files_frame,
                text="Обзор...",
                command=browse,
                bg="#ffcc80",
                activebackground="#ffb74d"
            )
            btn.grid(row=row, column=2, padx=5, pady=2)

        add_file_row(0, "Основной остаток (result_with_articul):", self.path_result)
        add_file_row(1, "Комиссия Kolobox (kolobox_commission_all_points):", self.path_commission)
        add_file_row(2, "Каталог Kolobox (catalog_tyres):", self.path_catalog)
        add_file_row(3, "Прайс 4tochki (Forto4ki_tires_detailed):", self.path_forto4ki)
        add_file_row(4, "Фиксированные цены (fixprice, опционально):", self.path_fixprice)

        ttk.Label(files_frame, text="Имя выходного файла:").grid(row=5, column=0, sticky="w")
        ttk.Entry(files_frame, textvariable=self.path_output).grid(row=5, column=1, sticky="ew", padx=5, pady=2)

        params_frame = ttk.LabelFrame(self, text="Параметры расчёта")
        params_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=5)
        for c in range(4):
            params_frame.columnconfigure(c, weight=1)

        self.var_min_margin = tk.DoubleVar(value=6.0)
        self.var_max_margin = tk.DoubleVar(value=15.0)
        self.var_low_qty = tk.IntVar(value=4)
        self.var_high_qty = tk.IntVar(value=20)
        self.var_popular_threshold = tk.IntVar(value=3)
        self.var_rare_threshold = tk.IntVar(value=2)
        self.var_popular_shift = tk.DoubleVar(value=-1.0)
        self.var_rare_shift = tk.DoubleVar(value=1.0)
        self.var_ox4_margin = tk.DoubleVar(value=10.0)
        self.var_address_ids = tk.StringVar(value="229,2139,2138,4344")

        ttk.Label(params_frame, text="Мин. маржа, %").grid(row=0, column=0, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_min_margin, width=7).grid(row=0, column=1, sticky="w")

        ttk.Label(params_frame, text="Макс. маржа, %").grid(row=0, column=2, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_max_margin, width=7).grid(row=0, column=3, sticky="w")

        ttk.Label(params_frame, text="Порог мал. кол-ва (шт.)").grid(row=1, column=0, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_low_qty, width=7).grid(row=1, column=1, sticky="w")

        ttk.Label(params_frame, text="Порог бол. кол-ва (шт.)").grid(row=1, column=2, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_high_qty, width=7).grid(row=1, column=3, sticky="w")

        ttk.Label(params_frame, text="Популярный размер: ≥ N вариантов").grid(row=2, column=0, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_popular_threshold, width=7).grid(row=2, column=1, sticky="w")

        ttk.Label(params_frame, text="Редкий размер: ≤ N вариантов").grid(row=2, column=2, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_rare_threshold, width=7).grid(row=2, column=3, sticky="w")

        ttk.Label(params_frame, text="Сдвиг маржи для популярных, %").grid(row=3, column=0, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_popular_shift, width=7).grid(row=3, column=1, sticky="w")

        ttk.Label(params_frame, text="Сдвиг маржи для редких/непопулярных, %").grid(row=3, column=2, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_rare_shift, width=7).grid(row=3, column=3, sticky="w")

        ttk.Label(params_frame, text="Маржа для ОХ-4, %").grid(row=4, column=0, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_ox4_margin, width=7).grid(row=4, column=1, sticky="w")

        ttk.Label(params_frame, text="ID наших складов (через запятую)").grid(row=4, column=2, sticky="w")
        ttk.Entry(params_frame, textvariable=self.var_address_ids, width=20).grid(row=4, column=3, sticky="w")

        pop_frame = ttk.LabelFrame(self, text="Популярные / непопулярные бренды и модели")
        pop_frame.grid(row=4, column=0, sticky="ew", padx=10, pady=5)
        for c in range(4):
            pop_frame.columnconfigure(c, weight=1)

        ttk.Label(pop_frame, text="Летние — популярные бренды:").grid(row=0, column=0, sticky="w")
        self.txt_summer_brands = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_summer_brands.grid(row=1, column=0, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Летние — популярные модели:").grid(row=0, column=1, sticky="w")
        self.txt_summer_models = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_summer_models.grid(row=1, column=1, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Зимние — популярные бренды:").grid(row=0, column=2, sticky="w")
        self.txt_winter_brands = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_winter_brands.grid(row=1, column=2, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Зимние — популярные модели:").grid(row=0, column=3, sticky="w")
        self.txt_winter_models = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_winter_models.grid(row=1, column=3, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Летние — НЕпопулярные бренды:").grid(row=2, column=0, sticky="w")
        self.txt_summer_brands_unpop = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_summer_brands_unpop.grid(row=3, column=0, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Летние — НЕпопулярные модели:").grid(row=2, column=1, sticky="w")
        self.txt_summer_models_unpop = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_summer_models_unpop.grid(row=3, column=1, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Зимние — НЕпопулярные бренды:").grid(row=2, column=2, sticky="w")
        self.txt_winter_brands_unpop = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_winter_brands_unpop.grid(row=3, column=2, sticky="nsew", padx=2, pady=2)

        ttk.Label(pop_frame, text="Зимние — НЕпопулярные модели:").grid(row=2, column=3, sticky="w")
        self.txt_winter_models_unpop = tk.Text(pop_frame, height=3, wrap="word")
        self.txt_winter_models_unpop.grid(row=3, column=3, sticky="nsew", padx=2, pady=2)

        run_frame = ttk.Frame(self)
        run_frame.grid(row=5, column=0, sticky="ew", padx=10, pady=5)
        run_frame.columnconfigure(0, weight=1)
        run_btn = tk.Button(
            run_frame,
            text="Запустить расчёт",
            command=self.on_run,
            bg="#ffcc80",
            activebackground="#ffb74d",
            font=("Segoe UI", 11, "bold")
        )
        run_btn.grid(row=0, column=0, pady=5)

        log_frame = ttk.LabelFrame(self, text="Лог")
        log_frame.grid(row=6, column=0, sticky="nsew", padx=10, pady=(0, 10))
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)

        self.log_text = tk.Text(log_frame, wrap="word")
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.configure(state="disabled")

        self.rowconfigure(6, weight=1)

    def toggle_instr(self):
        if self.instr_visible.get():
            self.instr_frame.grid_remove()
            self.instr_visible.set(False)
        else:
            self.instr_frame.grid()
            self.instr_visible.set(True)

    def on_run(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

        try:
            self._do_run()
        except Exception as e:
            self.log(f"ОШИБКА: {e}")
            messagebox.showerror("Ошибка", str(e))

    def _do_run(self):
        self.log("=== Старт расчёта ===")

        paths = {
            "result": self.path_result.get(),
            "commission": self.path_commission.get(),
            "catalog": self.path_catalog.get(),
            "forto4ki": self.path_forto4ki.get(),
        }
        optional_fix = self.path_fixprice.get()
        output_path = self.path_output.get()

        for key, p in paths.items():
            if not os.path.isfile(p):
                raise FileNotFoundError(f"Не найден файл: {p}")

        self.log("Все обязательные файлы найдены.")

        for label, p in paths.items():
            age = file_age_days(p)
            if age > 5:
                self.log(f"Предупреждение: файл {p} старше {age} дней.")

        self.log("Читаю основной остаток...")
        df_result_raw = pd.read_excel(paths["result"])

        self.log("Определяю сезоны по строке 'Зимние шины'...")
        df_result = determine_seasons(df_result_raw, log_func=self.log)

        self.log("Читаю комиссию Kolobox...")
        df_comm = pd.read_excel(paths["commission"])

        self.log("Читаю каталог Kolobox...")
        df_cat = pd.read_excel(paths["catalog"])

        self.log("Читаю прайс 4tochki (Forto4ki_tires_detailed, листы car/vned/cartruck)...")
        df_fort = load_forto4ki_detailed(paths["forto4ki"], sheets=('car', 'vned', 'cartruck'), log_func=self.log)

        df_fix = None
        if optional_fix and os.path.isfile(optional_fix):
            self.log("Читаю fixprice...")
            df_fix = pd.read_excel(optional_fix)
        else:
            self.log("Файл fixprice не найден или не указан — пропускаю фиксированные цены.")

        min_margin = float(self.var_min_margin.get()) / 100.0
        max_margin = float(self.var_max_margin.get()) / 100.0
        low_qty = int(self.var_low_qty.get())
        high_qty = int(self.var_high_qty.get())
        popular_thr = int(self.var_popular_threshold.get())
        rare_thr = int(self.var_rare_threshold.get())
        popular_shift = float(self.var_popular_shift.get()) / 100.0
        rare_shift = float(self.var_rare_shift.get()) / 100.0
        ox4_margin = float(self.var_ox4_margin.get()) / 100.0

        addr_str = self.var_address_ids.get()
        addr_ids = []
        for part in addr_str.split(","):
            part = part.strip()
            if part:
                try:
                    addr_ids.append(int(part))
                except ValueError:
                    self.log(f"Не удалось распарсить address_id '{part}', пропускаю его.")
        if not addr_ids:
            raise ValueError("Не заданы корректные ID складов.")

        summer_brands_text = self.txt_summer_brands.get("1.0", "end")
        summer_models_text = self.txt_summer_models.get("1.0", "end")
        winter_brands_text = self.txt_winter_brands.get("1.0", "end")
        winter_models_text = self.txt_winter_models.get("1.0", "end")

        summer_brands_unpop_text = self.txt_summer_brands_unpop.get("1.0", "end")
        summer_models_unpop_text = self.txt_summer_models_unpop.get("1.0", "end")
        winter_brands_unpop_text = self.txt_winter_brands_unpop.get("1.0", "end")
        winter_models_unpop_text = self.txt_winter_models_unpop.get("1.0", "end")

        summer_brand_keys = parse_keywords(summer_brands_text)
        summer_model_keys = parse_keywords(summer_models_text)
        winter_brand_keys = parse_keywords(winter_brands_text)
        winter_model_keys = parse_keywords(winter_models_text)

        summer_brand_keys_unpop = parse_keywords(summer_brands_unpop_text)
        summer_model_keys_unpop = parse_keywords(summer_models_unpop_text)
        winter_brand_keys_unpop = parse_keywords(winter_brands_unpop_text)
        winter_models_unpop = parse_keywords(winter_models_unpop_text)

        self.log(f"ID складов: {addr_ids}")
        self.log(f"Глобальная маржа: {min_margin*100:.1f}% – {max_margin*100:.1f}%")
        self.log(f"Маржа для ОХ-4: {ox4_margin*100:.1f}%")

        if (summer_brand_keys or summer_model_keys or winter_brand_keys or winter_model_keys or
                summer_brand_keys_unpop or summer_model_keys_unpop or
                winter_brand_keys_unpop or winter_models_unpop):
            self.log("Ключи популярных/непопулярных брендов/моделей заданы, будут применены соответствующие коэффициенты.")

        df_out, manual_needed_count = compute_auto_prices(
            df_result=df_result,
            commission_df=df_comm,
            catalog_df=df_cat,
            fort_df=df_fort,
            fixprice_df=df_fix,
            address_ids=tuple(addr_ids),
            global_min_margin=min_margin,
            global_max_margin=max_margin,
            low_qty=low_qty,
            high_qty=high_qty,
            popular_threshold=popular_thr,
            rare_threshold=rare_thr,
            popular_shift=popular_shift,
            rare_shift=rare_shift,
            alpha_max=0.7,
            ox4_margin_default=ox4_margin,
            summer_brand_keys=summer_brand_keys,
            summer_model_keys=summer_model_keys,
            winter_brand_keys=winter_brand_keys,
            winter_model_keys=winter_model_keys,
            summer_brand_keys_unpop=summer_brand_keys_unpop,
            summer_model_keys_unpop=summer_model_keys_unpop,
            winter_brand_keys_unpop=winter_brand_keys_unpop,
            winter_models_unpop=winter_models_unpop,
            log_func=self.log,
        )

        self.log("Сохраняю результат в Excel...")
        df_out.to_excel(output_path, index=False)
        self.log(f"Файл сохранён: {output_path}")

        color_special_rows_in_excel(output_path, log_func=self.log)

        kup_count = (df_out['Статус'].astype(str).str.lower() == 'купленный').sum()
        if kup_count > 0:
            self.log(f"Внимание: {kup_count} товаров со статусом 'Купленный' — задайте цены вручную или через fixprice.")

        if manual_needed_count > 0:
            self.log(
                f"Внимание: для {manual_needed_count} товаров нельзя подобрать цену в заданных рамках маржи — "
                f"они помечены как 'Нужна ручная цена' и оставлены без цены."
            )

        # Проверяем товары, которые есть в fixprice, но отсутствуют в текущем остатке
        if df_fix is not None and not df_fix.empty:
            fix_ids = set(df_fix.iloc[:, 0].astype(str))
            result_ids = set(df_result['Id'].astype(str))
            missing_fix_ids = sorted(fix_ids - result_ids)
            if missing_fix_ids:
                N = 50
                head_ids = ", ".join(missing_fix_ids[:N])
                self.log(
                    "Внимание: следующие артикула есть в fixprice, но отсутствуют в текущем остатке — "
                    "возможно, их пора удалить из fixprice:\n" + head_ids
                )
                if len(missing_fix_ids) > N:
                    self.log(f"... и ещё {len(missing_fix_ids) - N} шт.")

        self.log("=== Расчёт завершён ===")
        messagebox.showinfo("Готово", f"Расчёт завершён.\nРезультат: {output_path}")


if __name__ == "__main__":
    app = PricingApp()
    app.mainloop()
